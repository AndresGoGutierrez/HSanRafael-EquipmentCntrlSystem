from typing import List, Dict, Any, Optional
from io import BytesIO
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


class ReportGenerator:
    """Utility class for generating reports in PDF and Excel formats."""

    # === PDF REPORT ===
    @staticmethod
    def generate_pdf_report(
        title: str,
        data: List[Dict[str, Any]],
        headers: List[str],
        summary: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """
        Generate a PDF report.

        Args:
            title: Report title.
            data: List of data dictionaries.
            headers: Column headers.
            summary: Optional summary section with metrics.

        Returns:
            PDF file as bytes.
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []

        # === Styles ===
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=colors.HexColor("#1a1a1a"),
            spaceAfter=20,
            alignment=1,  # Center
        )

        # === Title ===
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.3 * inch))

        # === Summary section ===
        if summary:
            summary_data = [
                ["Metric", "Value"],
                ["Period Start", summary.get("period_start", "N/A")],
                ["Period End", summary.get("period_end", "N/A")],
                ["Total Entries", str(summary.get("total_entries", 0))],
                ["Total Exits", str(summary.get("total_exits", 0))],
                ["Currently Inside", str(summary.get("currently_inside", 0))],
                ["Expired Equipment", str(summary.get("expired_equipment", 0))],
            ]

            summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
            summary_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ]
                )
            )

            elements.extend([summary_table, Spacer(1, 0.3 * inch)])

        # === Data table ===
        if data:
            table_data = [headers] + [
                [str(row.get(h, "")) for h in headers] for row in data
            ]

            col_width = 6.5 * inch / len(headers)
            data_table = Table(table_data, colWidths=[col_width] * len(headers))
            data_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ]
                )
            )

            elements.append(data_table)
        else:
            elements.append(Paragraph("No data available.", styles["Normal"]))

        # === Build PDF ===
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # === EXCEL REPORT ===
    @staticmethod
    def generate_excel_report(
        title: str,
        data: List[Dict[str, Any]],
        headers: List[str],
        summary: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """
        Generate an Excel report.

        Args:
            title: Report title.
            data: List of data dictionaries.
            headers: Column headers.
            summary: Optional summary section with metrics.

        Returns:
            Excel file as bytes.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # === Title ===
        title_cell = ws.cell(row=1, column=1, value=title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        current_row = 3

        # === Summary section ===
        if summary:
            ws.cell(row=current_row, column=1, value="Summary").font = Font(
                bold=True, size=12
            )
            current_row += 1

            summary_items = [
                ("Period Start", str(summary.get("period_start", "N/A"))),
                ("Period End", str(summary.get("period_end", "N/A"))),
                ("Total Entries", summary.get("total_entries", 0)),
                ("Total Exits", summary.get("total_exits", 0)),
                ("Currently Inside", summary.get("currently_inside", 0)),
                ("Expired Equipment", summary.get("expired_equipment", 0)),
            ]

            for label, value in summary_items:
                ws.append([label, value])
                current_row += 1

            current_row += 2

        # === Headers ===
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        current_row += 1

        # === Data ===
        if data:
            for row_data in data:
                ws.append([str(row_data.get(h, "")) for h in headers])
        else:
            ws.append(["No data available."])

        # === Adjust column widths ===
        for column_cells in ws.columns:
            length = max(
                (len(str(cell.value or "")) for cell in column_cells), default=10
            )
            ws.column_dimensions[column_cells[0].column_letter].width = min(
                length + 2, 40
            )

        # === Save workbook ===
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
